import csv
import logging
from datetime import datetime, timedelta
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.views import LoginView
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Sum, Count
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.utils import timezone
from django.utils.http import urlencode
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from .forms import CustomLoginForm, CustomRegistrationForm
from .models import User, Category, Product, Cart, CartItem, Order, OrderItem, Review, Report
from django.db import transaction

logger = logging.getLogger(__name__)

def index(request):
    """Главная страница с категориями товаров."""
    categories = Category.objects.all()
    return render(request, 'flower_shop/index.html', {'categories': categories})

def catalog(request):
    """Страница каталога товаров."""
    products = Product.objects.filter(available=True)
    return render(request, 'flower_shop/catalog.html', {'products': products})

@login_required
def view_cart(request):
    """Страница корзины пользователя."""
    cart, created = Cart.objects.get_or_create(user=request.user)
    if created:
        logger.info(f"Корзина создана для пользователя {request.user.username}")
    return render(request, 'flower_shop/cart.html', {'cart': cart})

@login_required
def add_to_cart(request, product_id):
    """Добавление товара в корзину."""
    product = get_object_or_404(Product, id=product_id, available=True)
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_item, created = CartItem.objects.get_or_create(cart=cart, product=product)

    if not created:
        cart_item.quantity += 1
        cart_item.save()

    messages.success(request, f'Товар "{product.name}" добавлен в корзину.')
    return redirect(request.META.get('HTTP_REFERER', 'index'))

@login_required
def remove_from_cart(request, item_id):
    """Удаление товара из корзины."""
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    cart_item.delete()
    messages.success(request, 'Товар удален из корзины.')
    return redirect('cart_detail')


@login_required
def checkout(request):
    """Оформление заказа."""
    cart = get_object_or_404(Cart, user=request.user)
    if not cart.items.exists():
        messages.warning(request, 'Ваша корзина пуста.')
        logger.warning(f"Корзина пользователя {request.user.username} пуста.")
        return redirect('cart_detail')

    # Получаем последний заказ пользователя (если есть)
    last_order = Order.objects.filter(user=request.user).order_by('-created_at').first()

    # Значения по умолчанию
    default_name = request.user.email  # По умолчанию используем email пользователя
    default_phone = "+79991237567"  # Пример номера телефона
    default_address = "Самовывоз"  # Пример адреса
    default_comment = "Я люблю Лепесток"  # Пример комментария
    default_delivery_time = (timezone.now() + timedelta(days=1)).replace(hour=12, minute=0, second=0, microsecond=0)

    # Если есть последний заказ, используем его данные
    if last_order:
        default_name = last_order.name if last_order.name else request.user.email
        default_phone = last_order.phone if last_order.phone else default_phone
        default_address = last_order.address if last_order.address else default_address
        default_comment = last_order.comment if last_order.comment else default_comment

    # Если метод POST, значит пользователь либо нажал "Оформить заказ" в корзине, либо подтвердил заказ на странице checkout.html
    if request.method == 'POST':
        # Проверяем, откуда пришел POST-запрос
        if 'from_cart' in request.POST:  # Если запрос пришел из корзины
            # Создаем заказ со статусом "новый"
            with transaction.atomic():
                order = Order.objects.create(
                    user=request.user,
                    status='new',  # Статус "новый"
                    name=default_name,
                    phone=default_phone,
                    address=default_address,
                    delivery_time=default_delivery_time,
                    comment=default_comment
                )
                logger.info(f"Заказ #{order.id} создан со статусом 'новый'.")

                # Добавляем товары из корзины в заказ
                for item in cart.items.all():
                    OrderItem.objects.create(order=order, product=item.product, quantity=item.quantity)
                    logger.info(f"Товар {item.product.name} x {item.quantity} добавлен в заказ #{order.id}.")

            # Перенаправляем на страницу оформления заказа (checkout.html)
            return render(request, 'flower_shop/checkout.html', {
                'cart': cart,  # Передаем корзину в шаблон
                'default_name': default_name,
                'default_phone': default_phone,
                'default_address': default_address,
                'default_delivery_time': default_delivery_time.strftime('%Y-%m-%dT%H:%M'),
                'default_comment': default_comment,
                'telegram_id_attached': bool(request.user.telegram_id),
            })

        else:  # Если запрос пришел из checkout.html (подтверждение заказа)
            # Получаем данные из формы
            name = request.POST.get('name', default_name)
            phone = request.POST.get('phone', default_phone)
            address = request.POST.get('address', default_address)
            delivery_time = request.POST.get('delivery_time')
            comment = request.POST.get('comment', default_comment)

            # Преобразуем delivery_time в datetime
            if delivery_time:
                try:
                    naive_delivery_time = timezone.datetime.fromisoformat(delivery_time)
                    delivery_time = timezone.make_aware(naive_delivery_time, timezone.get_current_timezone())
                except (ValueError, TypeError):
                    delivery_time = default_delivery_time
            else:
                delivery_time = default_delivery_time

            # Используем транзакцию для обновления заказа
            with transaction.atomic():
                # Находим последний заказ пользователя со статусом "новый"
                order = Order.objects.filter(user=request.user, status='new').order_by('-created_at').first()
                if order:
                    # Обновляем заказ
                    order.status = 'accepted'  # Статус "принят"
                    order.name = name
                    order.phone = phone
                    order.address = address
                    order.delivery_time = delivery_time
                    order.comment = comment
                    order.save()
                    logger.info(f"Заказ #{order.id} обновлен со статусом 'принят'.")

                    # Очищаем корзину
                    cart.items.all().delete()
                    logger.info(f"Корзина пользователя {request.user.username} очищена.")

            messages.success(request, 'Заказ успешно оформлен!')
            return redirect('order_list')

    # Если метод GET, отображаем страницу оформления заказа
    return render(request, 'flower_shop/checkout.html', {
        'cart': cart,  # Передаем корзину в шаблон
        'default_name': default_name,
        'default_phone': default_phone,
        'default_address': default_address,
        'default_delivery_time': default_delivery_time.strftime('%Y-%m-%dT%H:%M'),
        'default_comment': default_comment,
        'telegram_id_attached': bool(request.user.telegram_id),
    })

@login_required
def order_list(request):
    """Список заказов с фильтрацией по статусам и датам."""
    selected_statuses = request.GET.getlist('status')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Для администраторов применяем фильтрацию по датам и статусам
    if request.user.is_staff:
        today_local = timezone.localtime(timezone.now()).date()
        if not start_date:
            start_date = today_local
        if not end_date:
            end_date = today_local

        # Преобразуем start_date и end_date в объекты date
        try:
            if isinstance(start_date, str):
                start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
            if isinstance(end_date, str):
                end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            start_date = today_local
            end_date = today_local

        start_datetime_utc = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
        end_datetime_utc = timezone.make_aware(timezone.datetime.combine(end_date, timezone.datetime.max.time()))
    else:
        # Для обычных пользователей не применяем фильтрацию
        start_datetime_utc = None
        end_datetime_utc = None
        selected_statuses = []

    if request.user.is_staff:
        orders = Order.objects.all().prefetch_related('items__product')
    else:
        orders = Order.objects.filter(user=request.user).prefetch_related('items__product')

    # Фильтрация по статусам и датам
    if request.user.is_staff and selected_statuses:
        orders = orders.filter(status__in=selected_statuses)
    if request.user.is_staff and start_datetime_utc and end_datetime_utc:
        orders = orders.filter(created_at__gte=start_datetime_utc, created_at__lte=end_datetime_utc)

    # Сортировка заказов по дате создания (сначала новые)
    orders = orders.order_by('-created_at')

    return render(request, 'flower_shop/order_list.html', {
        'orders': orders,
        'selected_statuses': selected_statuses,
        'status_choices': Order.STATUS_CHOICES,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
    })




def register(request):
    """Регистрация нового пользователя."""
    if request.method == 'POST':
        form = CustomRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            return redirect('index')
    else:
        form = CustomRegistrationForm()
    return render(request, 'flower_shop/register.html', {'form': form})


class CustomLoginView(LoginView):
    """Кастомный вход в систему."""
    form_class = CustomLoginForm
    template_name = 'flower_shop/login.html'

def product_list_by_category(request, category_slug):
    """Список товаров по категории."""
    category = get_object_or_404(Category, slug=category_slug)
    products_list = Product.objects.filter(category=category).order_by('name')
    paginator = Paginator(products_list, 6)
    page = request.GET.get('page')

    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)

    return render(request, 'flower_shop/product_list.html', {
        'category': category,
        'products': products,
    })

@login_required
def update_cart_item(request, item_id):
    """Обновление количества товара в корзине."""
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    if request.method == 'POST':
        try:
            quantity = int(request.POST.get('quantity', 1))
            if quantity > 0:
                cart_item.quantity = quantity
                cart_item.save()
            else:
                cart_item.delete()
        except (ValueError, TypeError):
            messages.error(request, 'Некорректное количество товара.')
    return redirect('cart_detail')

@login_required
def reorder(request, order_id):
    """Повторение заказа."""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    if not order.items.exists():
        messages.warning(request, 'Заказ пуст.')
        return redirect('order_list')

    cart, created = Cart.objects.get_or_create(user=request.user)

    for item in order.items.all():
        cart_item, created = CartItem.objects.get_or_create(cart=cart, product=item.product)
        if not created:
            cart_item.quantity += item.quantity
        else:
            cart_item.quantity = item.quantity
        cart_item.save()

    messages.success(request, 'Заказ добавлен в корзину.')
    return redirect('cart_detail')

def product_reviews(request, product_id):
    """Отзывы о товаре."""
    product = get_object_or_404(Product, id=product_id, available=True)
    reviews = Review.objects.filter(product=product)
    return render(request, 'flower_shop/product_reviews.html', {
        'product': product,
        'reviews': reviews,
    })

@login_required
def add_review(request, product_id):
    """Добавление отзыва."""
    product = get_object_or_404(Product, id=product_id, available=True)
    if request.method == 'POST':
        text = request.POST.get('text')
        rating = request.POST.get('rating')
        Review.objects.create(user=request.user, product=product, text=text, rating=rating)
        messages.success(request, 'Отзыв успешно добавлен.')
    return redirect('product_reviews', product_id=product.id)

@user_passes_test(lambda u: u.is_staff)
def change_order_status(request, order_id):
    """Изменение статуса заказа."""
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        new_status = request.POST.get('status')
        if new_status not in dict(Order.STATUS_CHOICES).keys():
            messages.error(request, 'Некорректный статус заказа.')
            return redirect('order_list')

        order.status = new_status
        order.save()

        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        filter_statuses = request.POST.getlist('filter_status')

        filter_params = {}
        if start_date:
            filter_params['start_date'] = start_date
        if end_date:
            filter_params['end_date'] = end_date
        if filter_statuses:
            filter_params['status'] = filter_statuses

        url = reverse('order_list')
        if filter_params:
            url += '?' + urlencode(filter_params, doseq=True)

        return redirect(url)
    return redirect('order_list')

@user_passes_test(lambda u: u.is_staff)
def order_report(request):
    """Отчёт по заказам."""
    start_date = request.GET.get('start_date', timezone.now().date())
    end_date = request.GET.get('end_date', timezone.now().date())

    if isinstance(start_date, str):
        start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
    if isinstance(end_date, str):
        end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()

    orders = Order.objects.filter(created_at__date__range=(start_date, end_date))

    # Убираем перенаправление на order_list
    if not orders.exists():
        messages.warning(request, 'Нет заказов за выбранный период.')
        # Вместо перенаправления, просто передаем пустые данные в шаблон
        return render(request, 'flower_shop/order_report.html', {
            'start_date': start_date,
            'end_date': end_date,
            'total_orders': 0,
            'total_revenue': 0,
            'daily_stats': [],
        })

    daily_stats = (
        orders
        .values('created_at__date')
        .annotate(
            total_orders=Count('id'),
            total_revenue=Sum('items__product__price')
        )
        .order_by('created_at__date')
    )

    total_orders = orders.count()
    total_revenue = orders.aggregate(total_revenue=Sum('items__product__price'))['total_revenue'] or 0

    return render(request, 'flower_shop/order_report.html', {
        'start_date': start_date,
        'end_date': end_date,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'daily_stats': daily_stats,
    })

@user_passes_test(lambda u: u.is_staff)
def export_order_report(request):
    """Экспорт отчёта в CSV."""
    start_date = request.GET.get('start_date', timezone.now().date())
    end_date = request.GET.get('end_date', timezone.now().date())
    orders = Order.objects.filter(created_at__date__range=(start_date, end_date))
    if not orders.exists():
        messages.warning(request, 'Нет заказов за выбранный период.')
        return redirect('order_list')

    total_orders = orders.count()
    total_revenue = sum(order.get_total_price() for order in orders)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="order_report_{start_date}_{end_date}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Период', 'Количество заказов', 'Выручка'])
    writer.writerow([f'{start_date} - {end_date}', total_orders, total_revenue])

    return response

@user_passes_test(lambda u: u.is_staff)
def export_order_report_excel(request):
    """Экспорт отчёта в Excel."""
    start_date = request.GET.get('start_date', timezone.now().date())
    end_date = request.GET.get('end_date', timezone.now().date())
    orders = Order.objects.filter(created_at__date__range=(start_date, end_date))
    if not orders.exists():
        messages.warning(request, 'Нет заказов за выбранный период.')
        return redirect('order_list')

    total_orders = orders.count()
    total_revenue = sum(order.get_total_price() for order in orders)

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт по заказам"

    ws['A1'] = "Отчёт по заказам"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:C1')

    ws['A2'] = f"Период: {start_date} - {end_date}"
    ws['A2'].font = Font(bold=True)
    ws.merge_cells('A2:C2')

    ws['A3'] = f"Общее количество заказов: {total_orders}"
    ws['A4'] = f"Общая выручка: {total_revenue} руб."
    ws['A3'].font = Font(bold=True)
    ws['A4'].font = Font(bold=True)

    ws['A6'] = "Дата"
    ws['B6'] = "Количество заказов"
    ws['C6'] = "Выручка"
    for cell in ws['6:6']:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    daily_stats = (
        orders
        .values('created_at__date')
        .annotate(
            total_orders=Count('id'),
            total_revenue=Sum('items__product__price')
        )
        .order_by('created_at__date')
    )

    row = 7
    for stat in daily_stats:
        ws[f'A{row}'] = stat['created_at__date'].strftime('%d.%m.%Y')
        ws[f'B{row}'] = stat['total_orders']
        ws[f'C{row}'] = stat['total_revenue']
        row += 1

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="order_report_{start_date}_{end_date}.xlsx"'
    wb.save(response)

    return response

@login_required
def generate_link_code(request):
    """Генерация кода для привязки Telegram."""
    try:
        user = request.user
        code = user.generate_link_code()
        logger.info(f"User {user.username} generated a link code: {code}")
        return JsonResponse({'code': code})
    except Exception as e:
        logger.error(f"Error generating link code for user {user.username}: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)

def auth_with_telegram_id(request):
    """Авторизация пользователя по Telegram ID."""
    telegram_id = request.GET.get('telegram_id')
    if not telegram_id:
        return redirect('index')

    try:
        user = User.objects.get(telegram_id=telegram_id)
        login(request, user, backend='django.contrib.auth.backends.ModelBackend')
        return redirect('index')
    except User.DoesNotExist:
        messages.error(request, 'Пользователь с таким Telegram ID не найден.')
        return redirect('index')